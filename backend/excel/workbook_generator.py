"""
Nuvel — 5-sheet Excel workbook generator.

Sheet layout (row numbers are 1-based, used in cross-sheet formula refs):

  Sheet 1 "Raw Data"
    Row 1  : Column headers
    Row 2  : ── INCOME STATEMENT ──
    Row 3  : Total Revenue            (totalRevenue)
    Row 4  : Cost of Revenue          (derived: revenue - grossProfit)
    Row 5  : Gross Profit             (grossProfit)
    Row 6  : R&D Expense              (researchAndDevelopment)
    Row 7  : SG&A Expense             (sellingGeneralAndAdministrative)
    Row 8  : Operating Income         (operatingIncome)
    Row 9  : EBITDA                   (ebitda)
    Row 10 : Net Income               (netIncome)
    Row 11 : Diluted EPS              (eps)
    Row 12 : blank
    Row 13 : ── BALANCE SHEET ──
    Row 14 : Cash & Equivalents       (cashAndCashEquivalentsAtCarryingValue)
    Row 15 : Net Receivables          (currentNetReceivables)
    Row 16 : Inventory                (inventory)
    Row 17 : Total Current Assets     (totalCurrentAssets)
    Row 18 : Total Assets             (totalAssets)
    Row 19 : Total Current Liabilities(totalCurrentLiabilities)
    Row 20 : Long-Term Debt           (longTermDebt)
    Row 21 : Shareholder Equity       (totalShareholderEquity)
    Row 22 : Goodwill                 (goodwill)
    Row 23 : blank
    Row 24 : ── CASH FLOW ──
    Row 25 : Operating Cash Flow      (operatingCashflow)
    Row 26 : Capital Expenditures     (capitalExpenditures)
    Row 27 : Free Cash Flow           (freeCashFlow)

  Data columns: A=CN name, B=EN name, C=oldest yr … G=newest yr,
                H=YoY%, I=Source, J=Page ref
  Most-recent-year column = G  (col index 7)

  Sheet 2 "Ratios"  — formulas reference 'Raw Data'!G<row>

  Sheet 3 "DCF Model"
    B3:B11  = yellow assumption cells
    B15:K26 = 10-year projection table
    B28:B35 = terminal value + valuation
    B37:D39 = scenario table (bear / base / bull)
"""

from __future__ import annotations

import io
import logging
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Palette ────────────────────────────────────────────────────────────────
_BLUE   = "4472C4"
_GREEN  = "70AD47"
_ORANGE = "ED7D31"
_RED    = "C00000"
_YELLOW = "FFF2CC"
_LIGHT_BLUE  = "DEEAF1"
_LIGHT_GREEN = "E2EFDA"
_LIGHT_RED   = "FCE4D6"
_GRAY   = "F2F2F2"
_WHITE  = "FFFFFF"

# ── Style factories ────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold: bool = False, color: str = "000000", size: int = 11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")

def _thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _freeze(ws, cell: str = "A2") -> None:
    ws.freeze_panes = ws[cell]

def _autofit(ws, min_width: int = 10, max_width: int = 40) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        best = max(
            (len(str(c.value or "")) for c in col if c.value is not None),
            default=min_width,
        )
        ws.column_dimensions[col_letter].width = max(min_width, min(best + 2, max_width))

def _section_header(ws, row: int, label: str, n_cols: int = 10) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=label)
    cell.fill  = _fill(_GRAY)
    cell.font  = _font(bold=True, size=10)
    cell.alignment = _left()

# ── Sheet 1: Raw Data ──────────────────────────────────────────────────────

# (row, CN name, EN name, income_col, balance_col, cashflow_col, source_label)
_RAW_ROWS: list[tuple] = [
    # Section: Income Statement
    (3,  "总营收",       "Total Revenue",            "totalRevenue",            None, None, "AV INCOME_STATEMENT"),
    (4,  "营业成本",     "Cost of Revenue",          "__cogs__",                None, None, "AV INCOME_STATEMENT"),
    (5,  "毛利润",       "Gross Profit",             "grossProfit",             None, None, "AV INCOME_STATEMENT"),
    (6,  "研发费用",     "R&D Expense",              "researchAndDevelopment",  None, None, "AV INCOME_STATEMENT"),
    (7,  "销售及管理费用","SG&A Expense",             "sellingGeneralAndAdministrative", None, None, "AV INCOME_STATEMENT"),
    (8,  "营业利润",     "Operating Income",         "operatingIncome",         None, None, "AV INCOME_STATEMENT"),
    (9,  "息税折旧前利润","EBITDA",                  "ebitda",                  None, None, "AV INCOME_STATEMENT"),
    (10, "净利润",       "Net Income",               "netIncome",               None, None, "AV INCOME_STATEMENT"),
    (11, "每股收益（摊薄）","Diluted EPS",            "eps",                     None, None, "AV INCOME_STATEMENT"),
    # Section: Balance Sheet
    (14, "现金及等价物",  "Cash & Equivalents",       None, "cashAndCashEquivalentsAtCarryingValue", None, "AV BALANCE_SHEET"),
    (15, "应收账款净额",  "Net Receivables",          None, "currentNetReceivables",                 None, "AV BALANCE_SHEET"),
    (16, "存货",          "Inventory",               None, "inventory",                             None, "AV BALANCE_SHEET"),
    (17, "流动资产合计",  "Total Current Assets",    None, "totalCurrentAssets",                    None, "AV BALANCE_SHEET"),
    (18, "总资产",        "Total Assets",            None, "totalAssets",                           None, "AV BALANCE_SHEET"),
    (19, "流动负债合计",  "Total Current Liabilities",None,"totalCurrentLiabilities",               None, "AV BALANCE_SHEET"),
    (20, "长期债务",      "Long-Term Debt",          None, "longTermDebt",                          None, "AV BALANCE_SHEET"),
    (21, "股东权益",      "Shareholder Equity",      None, "totalShareholderEquity",                None, "AV BALANCE_SHEET"),
    (22, "商誉",          "Goodwill",                None, "goodwill",                              None, "AV BALANCE_SHEET"),
    # Section: Cash Flow
    (25, "经营现金流",    "Operating Cash Flow",     None, None, "operatingCashflow",    "AV CASH_FLOW"),
    (26, "资本支出",      "Capital Expenditures",    None, None, "capitalExpenditures",  "AV CASH_FLOW"),
    (27, "自由现金流",    "Free Cash Flow",          None, None, "freeCashFlow",         "AV CASH_FLOW"),
]

# Most-recent-year data column letter (G = column 7)
_DATA_COL = 7   # 1-based column index for the newest year in Sheet 1


def _build_raw_data(ws, income: pd.DataFrame, balance: pd.DataFrame, cashflow: pd.DataFrame) -> None:
    ws.title = "Raw Data"
    ws.sheet_properties.tabColor = _BLUE

    # ── Header row ─────────────────────────────────────────────────────────
    headers = ["科目（中文）", "Account (English)"]
    years: list[str] = []

    # Collect up to 5 fiscal years from income DF (newest first → reverse)
    for df in (income, balance, cashflow):
        if df is not None and "fiscalDateEnding" in df.columns:
            for ts in reversed(df["fiscalDateEnding"].tolist()):
                yr = str(pd.Timestamp(ts).year)
                if yr not in years:
                    years.append(yr)
            break

    # Pad to exactly 5 slots (oldest in C, newest in G)
    while len(years) < 5:
        years.insert(0, "N/A")
    years = years[-5:]

    headers += years + ["YoY %", "Data Source", "Page Ref"]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _fill(_BLUE)
        cell.font = _font(bold=True, color="FFFFFF")
        cell.alignment = _center()

    # ── Section headers ────────────────────────────────────────────────────
    _section_header(ws, 2,  "── INCOME STATEMENT ──")
    _section_header(ws, 13, "── BALANCE SHEET ──")
    _section_header(ws, 24, "── CASH FLOW STATEMENT ──")

    # ── Helper: get value from the right DataFrame ─────────────────────────
    def _val(income_col, balance_col, cashflow_col, df_idx: int):
        """Retrieve a cell value; df_idx 0=oldest, 4=newest (of available rows)."""
        for col_name, df in (
            (income_col, income), (balance_col, balance), (cashflow_col, cashflow)
        ):
            if col_name and df is not None:
                if col_name == "__cogs__":
                    # Derived: Cost of Revenue = Revenue - Gross Profit
                    try:
                        rev = pd.to_numeric(df.iloc[df_idx].get("totalRevenue"), errors="coerce")
                        gp  = pd.to_numeric(df.iloc[df_idx].get("grossProfit"),  errors="coerce")
                        return int(rev - gp) if pd.notna(rev) and pd.notna(gp) else None
                    except IndexError:
                        return None
                try:
                    v = df.iloc[df_idx].get(col_name)
                    v = pd.to_numeric(v, errors="coerce")
                    return None if pd.isna(v) else int(v)
                except IndexError:
                    return None
        return None

    # Column layout: C=oldest … G=newest (right-aligned).
    # AlphaVantage DataFrame: row 0 = newest, row 4 = oldest.
    # slot 0 → col C (oldest display), slot 4 → col G (newest display).
    # df_idx for slot s = 4 - s  (slot 4 → df_idx 0 = newest ✓)
    # Valid only when df_idx < n_avail.
    n_avail = len(income) if income is not None else 0

    for (row, cn, en, ic, bc, cc, src) in _RAW_ROWS:
        ws.cell(row=row, column=1, value=cn).alignment = _left()
        ws.cell(row=row, column=2, value=en).alignment = _left()

        for slot in range(5):           # 0=oldest col C … 4=newest col G
            col_idx = slot + 3          # C=3, D=4, E=5, F=6, G=7
            df_idx  = 4 - slot          # slot 0→df_idx 4 (oldest), slot 4→df_idx 0 (newest)
            if df_idx >= n_avail:       # year not available → dash
                ws.cell(row=row, column=col_idx, value="—").alignment = _center()
            else:
                v = _val(ic, bc, cc, df_idx)
                cell = ws.cell(row=row, column=col_idx, value=v)
                cell.number_format = "#,##0"
                cell.alignment = _center()

        # YoY% in col H: (newest - prior) / |prior|
        newest   = _val(ic, bc, cc, 0)
        prior    = _val(ic, bc, cc, 1) if n_avail > 1 else None
        if newest is not None and prior is not None and prior != 0:
            yoy = (newest - prior) / abs(prior)
            h_cell = ws.cell(row=row, column=8, value=yoy)
            h_cell.number_format = "0.0%"
            h_cell.alignment = _center()
        else:
            ws.cell(row=row, column=8, value="N/A").alignment = _center()

        ws.cell(row=row, column=9, value=src).alignment = _left()
        ws.cell(row=row, column=10, value="SEC EDGAR / AV API").alignment = _left()

    # ── Conditional formatting on YoY column ──────────────────────────────
    ws.conditional_formatting.add(
        "H3:H27",
        CellIsRule(operator="greaterThan", formula=["0"], fill=_fill("C6EFCE"),
                   font=Font(color="276221")),
    )
    ws.conditional_formatting.add(
        "H3:H27",
        CellIsRule(operator="lessThan", formula=["0"], fill=_fill("FFC7CE"),
                   font=Font(color="9C0006")),
    )

    _freeze(ws)
    _autofit(ws)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["I"].width = 26
    ws.column_dimensions["J"].width = 24


# ── Sheet 2: Ratios ────────────────────────────────────────────────────────

_RATIO_DEFS: list[dict] = [
    # Profitability
    {"group": "Profitability", "name_cn": "毛利率", "name_en": "Gross Margin",
     "formula": "='Raw Data'!G5/'Raw Data'!G3", "fmt": "0.0%",
     "comment": "Gross Profit / Revenue. Typical SaaS/semis: 50-80%.", "industry": 0.55},
    {"group": "Profitability", "name_cn": "营业利润率", "name_en": "Operating Margin",
     "formula": "='Raw Data'!G8/'Raw Data'!G3", "fmt": "0.0%",
     "comment": "Operating Income / Revenue. Healthy: >15%.", "industry": 0.20},
    {"group": "Profitability", "name_cn": "净利润率", "name_en": "Net Margin",
     "formula": "='Raw Data'!G10/'Raw Data'!G3", "fmt": "0.0%",
     "comment": "Net Income / Revenue. Industry avg: 10-20%.", "industry": 0.15},
    {"group": "Profitability", "name_cn": "资产收益率", "name_en": "Return on Assets",
     "formula": "='Raw Data'!G10/'Raw Data'!G18", "fmt": "0.0%",
     "comment": "Net Income / Total Assets. Healthy: >5%.", "industry": 0.08},
    {"group": "Profitability", "name_cn": "股本回报率", "name_en": "Return on Equity",
     "formula": "='Raw Data'!G10/'Raw Data'!G21", "fmt": "0.0%",
     "comment": "Net Income / Equity. Healthy: >15%.", "industry": 0.18},
    # Liquidity
    {"group": "Liquidity", "name_cn": "流动比率", "name_en": "Current Ratio",
     "formula": "='Raw Data'!G17/'Raw Data'!G19", "fmt": "0.00x",
     "comment": "Current Assets / Current Liabilities. Target: ≥1.5x.", "industry": 1.80},
    {"group": "Liquidity", "name_cn": "现金比率", "name_en": "Cash Ratio",
     "formula": "='Raw Data'!G14/'Raw Data'!G19", "fmt": "0.00x",
     "comment": "Cash / Current Liabilities. Conservative liquidity measure.", "industry": 0.50},
    # Leverage
    {"group": "Leverage", "name_cn": "债务权益比", "name_en": "Debt-to-Equity",
     "formula": "='Raw Data'!G20/'Raw Data'!G21", "fmt": "0.00x",
     "comment": "Long-Term Debt / Equity. Low is safer; <1x preferred.", "industry": 0.60},
    # Efficiency
    {"group": "Efficiency", "name_cn": "资产周转率", "name_en": "Asset Turnover",
     "formula": "='Raw Data'!G3/'Raw Data'!G18", "fmt": "0.00x",
     "comment": "Revenue / Total Assets. Higher = more efficient use of assets.", "industry": 0.65},
    {"group": "Efficiency", "name_cn": "应收账款/营收", "name_en": "Receivables / Revenue",
     "formula": "='Raw Data'!G15/'Raw Data'!G3", "fmt": "0.0%",
     "comment": "High ratio may signal collection risk.", "industry": 0.12},
    # R&D & Spending
    {"group": "R&D & Spending", "name_cn": "研发强度", "name_en": "R&D / Revenue",
     "formula": "='Raw Data'!G6/'Raw Data'!G3", "fmt": "0.0%",
     "comment": "R&D spend as % of revenue. Semis avg: 10-20%.", "industry": 0.15},
    {"group": "R&D & Spending", "name_cn": "销管费率", "name_en": "SG&A / Revenue",
     "formula": "='Raw Data'!G7/'Raw Data'!G3", "fmt": "0.0%",
     "comment": "SG&A efficiency. Lower is better for mature companies.", "industry": 0.10},
    # Cash Flow
    {"group": "Cash Flow", "name_cn": "自由现金流", "name_en": "Free Cash Flow ($)",
     "formula": "='Raw Data'!G25-'Raw Data'!G26", "fmt": '#,##0',
     "comment": "Operating CF - CapEx. Absolute FCF generation.", "industry": None},
    {"group": "Cash Flow", "name_cn": "资本支出/营收", "name_en": "CapEx / Revenue",
     "formula": "='Raw Data'!G26/'Raw Data'!G3", "fmt": "0.0%",
     "comment": "Capital intensity. Asset-light: <5%, heavy: >15%.", "industry": 0.06},
    # Quality
    {"group": "Quality", "name_cn": "商誉/总资产", "name_en": "Goodwill / Assets",
     "formula": "='Raw Data'!G22/'Raw Data'!G18", "fmt": "0.0%",
     "comment": "High ratio (>30%) signals acquisition-heavy growth.", "industry": 0.15},
]


def _build_ratios(ws, ratios: dict) -> None:
    ws.title = "Ratios"
    ws.sheet_properties.tabColor = _GREEN

    # Header
    for col, h in enumerate(["类别", "指标（中文）", "Ratio (EN)", "Value", "YoY %", "Industry Avg", "Comment"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _fill(_GREEN)
        cell.font = _font(bold=True, color="FFFFFF")
        cell.alignment = _center()

    current_group = None
    row = 2
    for rd in _RATIO_DEFS:
        if rd["group"] != current_group:
            current_group = rd["group"]
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            gc = ws.cell(row=row, column=1, value=f"— {current_group} —")
            gc.fill = _fill(_GRAY)
            gc.font = _font(bold=True, size=10)
            gc.alignment = _left()
            row += 1

        ws.cell(row=row, column=1, value=rd["group"]).alignment = _center()
        ws.cell(row=row, column=2, value=rd["name_cn"]).alignment = _left()
        ws.cell(row=row, column=3, value=rd["name_en"]).alignment = _left()

        # Formula cell — the actual cross-sheet reference
        vcell = ws.cell(row=row, column=4)
        vcell.value = rd["formula"]
        vcell.number_format = rd["fmt"]
        vcell.alignment = _center()

        # YoY placeholder (no prior-year ratio stored — leave for future expansion)
        ws.cell(row=row, column=5, value="—").alignment = _center()

        # Industry average
        if rd["industry"] is not None:
            ic = ws.cell(row=row, column=6, value=rd["industry"])
            ic.number_format = rd["fmt"]
            ic.alignment = _center()
            ic.fill = _fill(_LIGHT_BLUE)
        else:
            ws.cell(row=row, column=6, value="—").alignment = _center()

        # Cell comment
        comment = Comment(rd["comment"], "Nuvel")
        comment.width = 250
        comment.height = 80
        ws.cell(row=row, column=3).comment = comment

        row += 1

    _freeze(ws)
    _autofit(ws)


# ── Sheet 3: DCF Model ─────────────────────────────────────────────────────

_ASSUMPTION_LABELS = [
    ("营收增长率（第1-5年）",  "Revenue CAGR Y1–5",  "0.0%"),
    ("营收增长率（第6-10年）", "Revenue CAGR Y6–10", "0.0%"),
    ("永续增长率",             "Terminal Growth Rate","0.0%"),
    ("EBITDA利润率",          "EBITDA Margin",       "0.0%"),
    ("有效税率",               "Effective Tax Rate",  "0.0%"),
    ("资本支出/营收",         "CapEx % Revenue",     "0.0%"),
    ("折旧摊销/营收",         "D&A % Revenue",       "0.0%"),
    ("营运资本变动/营收",     "NWC Change % Revenue","0.0%"),
    ("加权平均资本成本",      "WACC",                "0.0%"),
]


def _build_dcf(ws, income: pd.DataFrame, balance: pd.DataFrame) -> None:
    ws.title = "DCF Model"
    ws.sheet_properties.tabColor = _ORANGE

    # Title
    ws.merge_cells("A1:L1")
    t = ws["A1"]
    t.value = "DCF VALUATION MODEL  —  Assumptions highlighted in yellow"
    t.font = _font(bold=True, size=13, color=_ORANGE)
    t.alignment = _left()

    # ── Assumption inputs (B3:B11, yellow fill) ────────────────────────────
    ws["A2"] = "Assumptions"
    ws["A2"].font = _font(bold=True)

    # Derive defaults from actual data
    def _pct(col, df=income) -> float:
        try:
            v = pd.to_numeric(df.iloc[0].get(col, 0), errors="coerce")
            r = pd.to_numeric(df.iloc[0].get("totalRevenue", 1), errors="coerce")
            return round(float(v / r), 4) if r and r != 0 else 0.0
        except Exception:
            return 0.0

    defaults = [
        0.20,                          # Revenue CAGR Y1-5  (placeholder)
        0.12,                          # Revenue CAGR Y6-10
        0.025,                         # Terminal growth
        _pct("ebitda"),                # EBITDA margin from data
        0.21,                          # Tax rate
        _pct("capitalExpenditures", income) if False else 0.03,  # CapEx (use cashflow later)
        0.04,                          # D&A % revenue
        0.02,                          # NWC change
        0.10,                          # WACC
    ]

    yellow_fill = _fill(_YELLOW)
    assumption_rows: dict[str, int] = {}   # label → row number (for formula refs)

    for i, (cn, en, fmt) in enumerate(_ASSUMPTION_LABELS):
        r = i + 3
        ws.cell(row=r, column=1, value=cn).alignment = _left()
        ws.cell(row=r, column=2, value=en).font = _font(bold=True)
        vc = ws.cell(row=r, column=3, value=defaults[i])
        vc.fill = yellow_fill
        vc.number_format = fmt
        vc.alignment = _center()
        assumption_rows[en] = r

    # Short cell refs for assumption row numbers (1-based)
    rG1   = assumption_rows["Revenue CAGR Y1–5"]
    rG2   = assumption_rows["Revenue CAGR Y6–10"]
    rTGR  = assumption_rows["Terminal Growth Rate"]
    rEBIT = assumption_rows["EBITDA Margin"]
    rTAX  = assumption_rows["Effective Tax Rate"]
    rCX   = assumption_rows["CapEx % Revenue"]
    rDA   = assumption_rows["D&A % Revenue"]
    rNWC  = assumption_rows["NWC Change % Revenue"]
    rWACC = assumption_rows["WACC"]

    # ── Projection table (rows 14-26) ─────────────────────────────────────
    ws["A13"] = "10-Year Projection  (values in $000s)"
    ws["A13"].font = _font(bold=True)

    proj_start = 14
    proj_labels = [
        "Revenue",
        "EBITDA",
        "EBIT (after D&A)",
        "NOPAT",
        "Plus: D&A",
        "Less: CapEx",
        "Less: ΔNWC",
        "FCFF",
        "Discount Factor",
        "PV of FCFF",
    ]
    for i, lbl in enumerate(proj_labels):
        ws.cell(row=proj_start + i, column=1, value=lbl).font = _font(bold=(i == 7))

    # Year headers (cols D=4 … M=13)
    base_rev_ref = "='Raw Data'!G3"  # most recent revenue from Sheet 1
    for yr_idx in range(10):
        col = yr_idx + 4  # D through M
        col_letter = get_column_letter(col)
        prev_col   = get_column_letter(col - 1)

        yr_label = f"Year {yr_idx + 1}"
        ws.cell(row=proj_start - 1, column=col, value=yr_label)
        ws.cell(row=proj_start - 1, column=col).font = _font(bold=True, color="FFFFFF")
        ws.cell(row=proj_start - 1, column=col).fill = _fill(_ORANGE)
        ws.cell(row=proj_start - 1, column=col).alignment = _center()

        r_rev  = proj_start
        r_ebit = proj_start + 1
        r_ebi2 = proj_start + 2
        r_nopa = proj_start + 3
        r_da   = proj_start + 4
        r_cx   = proj_start + 5
        r_nwc  = proj_start + 6
        r_fcff = proj_start + 7
        r_df   = proj_start + 8
        r_pv   = proj_start + 9

        growth_ref = f"$C${rG1}" if yr_idx < 5 else f"$C${rG2}"

        if yr_idx == 0:
            rev_formula = f"='Raw Data'!G3*(1+{growth_ref})"
        else:
            rev_formula = f"={prev_col}{r_rev}*(1+{growth_ref})"

        ws.cell(row=r_rev,  column=col).value = rev_formula
        ws.cell(row=r_ebit, column=col).value = f"={col_letter}{r_rev}*$C${rEBIT}"
        ws.cell(row=r_ebi2, column=col).value = f"={col_letter}{r_ebit}-{col_letter}{r_rev}*$C${rDA}"
        ws.cell(row=r_nopa, column=col).value = f"={col_letter}{r_ebi2}*(1-$C${rTAX})"
        ws.cell(row=r_da,   column=col).value = f"={col_letter}{r_rev}*$C${rDA}"
        ws.cell(row=r_cx,   column=col).value = f"={col_letter}{r_rev}*$C${rCX}"
        ws.cell(row=r_nwc,  column=col).value = f"={col_letter}{r_rev}*$C${rNWC}"
        ws.cell(row=r_fcff, column=col).value = (
            f"={col_letter}{r_nopa}+{col_letter}{r_da}"
            f"-{col_letter}{r_cx}-{col_letter}{r_nwc}"
        )
        disc_exp = yr_idx + 1
        ws.cell(row=r_df,   column=col).value = f"=1/(1+$C${rWACC})^{disc_exp}"
        ws.cell(row=r_pv,   column=col).value = f"={col_letter}{r_fcff}*{col_letter}{r_df}"

        for r in range(proj_start, proj_start + 10):
            ws.cell(row=r, column=col).number_format = "#,##0"
            ws.cell(row=r, column=col).alignment = _center()
        ws.cell(row=r_df, column=col).number_format = "0.0000"

    # ── Terminal value & valuation (rows 29-36) ───────────────────────────
    last_col = "M"   # Year 10
    tv_start = proj_start + 12

    ws.cell(row=tv_start,   column=1, value="Terminal FCF").font = _font(bold=True)
    ws.cell(row=tv_start+1, column=1, value="Terminal Value").font = _font(bold=True)
    ws.cell(row=tv_start+2, column=1, value="PV of Terminal Value").font = _font(bold=True)
    ws.cell(row=tv_start+3, column=1, value="PV of FCFFs (Sum)").font = _font(bold=True)
    ws.cell(row=tv_start+4, column=1, value="Enterprise Value").font = _font(bold=True)
    ws.cell(row=tv_start+5, column=1, value="Net Debt (Long-term)").font = _font(bold=True)
    ws.cell(row=tv_start+6, column=1, value="Equity Value").font = _font(bold=True)
    ws.cell(row=tv_start+7, column=1, value="Shares Outstanding (M)").font = _font(bold=True)
    ws.cell(row=tv_start+8, column=1, value="Intrinsic Value / Share").font = _font(bold=True, color=_ORANGE)

    r_fcff_last = proj_start + 7
    r_df_last   = proj_start + 8

    ws["C" + str(tv_start)]   = f"={last_col}{r_fcff_last}*(1+$C${rTGR})"
    ws["C" + str(tv_start+1)] = f"=C{tv_start}/($C${rWACC}-$C${rTGR})"
    ws["C" + str(tv_start+2)] = f"=C{tv_start+1}*{last_col}{r_df_last}"
    ws["C" + str(tv_start+3)] = (
        f"=SUM(D{proj_start+9}:M{proj_start+9})"
    )
    ws["C" + str(tv_start+4)] = f"=C{tv_start+3}+C{tv_start+2}"
    ws["C" + str(tv_start+5)] = "='Raw Data'!G20"   # Long-term debt from Sheet 1
    ws["C" + str(tv_start+6)] = f"=C{tv_start+4}-C{tv_start+5}"
    ws.cell(row=tv_start+7, column=3, value=24400)   # placeholder shares (NVDA ~24.4B)
    ws["C" + str(tv_start+8)] = f"=C{tv_start+6}/C{tv_start+7}/1000"

    for r in range(tv_start, tv_start + 9):
        ws.cell(row=r, column=3).number_format = "#,##0"
        ws.cell(row=r, column=3).alignment = _center()
    ws.cell(row=tv_start + 8, column=3).number_format = "$#,##0.00"

    # ── Scenario table ────────────────────────────────────────────────────
    sc_row = tv_start + 11
    ws.cell(row=sc_row, column=1, value="Scenario Analysis").font = _font(bold=True, size=12)
    for col, h in enumerate(["Scenario", "Growth Y1-5", "WACC", "Intrinsic Value/Share"], 1):
        ws.cell(row=sc_row + 1, column=col, value=h).font = _font(bold=True, color="FFFFFF")
        ws.cell(row=sc_row + 1, column=col).fill = _fill(_ORANGE)
        ws.cell(row=sc_row + 1, column=col).alignment = _center()

    for i, (label, g_adj, w_adj) in enumerate([("🐻 Bear", -0.05, 0.01), ("📊 Base", 0.0, 0.0), ("🐂 Bull", 0.05, -0.01)]):
        r = sc_row + 2 + i
        ws.cell(row=r, column=1, value=label).alignment = _center()
        ws.cell(row=r, column=2, value=f"=$C${rG1}+{g_adj}").number_format = "0.0%"
        ws.cell(row=r, column=2).alignment = _center()
        ws.cell(row=r, column=3, value=f"=$C${rWACC}+{w_adj}").number_format = "0.0%"
        ws.cell(row=r, column=3).alignment = _center()
        ws.cell(row=r, column=4, value="See base model →").alignment = _center()

    _freeze(ws)
    _autofit(ws)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 26


# ── Sheet 4: Expectations ──────────────────────────────────────────────────

def _build_expectations(ws, ticker: str, year: str) -> None:
    ws.title = "Expectations"

    ws.merge_cells("A1:F1")
    h = ws["A1"]
    h.value = f"Analyst Expectations — {ticker.upper()} {year}  (Source: Yahoo Finance)"
    h.font = _font(bold=True, size=12)
    h.alignment = _left()

    headers = ["Metric", "Consensus Est.", "High Est.", "Low Est.", "# Analysts", "Surprise %"]
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.fill = _fill(_GRAY)
        cell.font = _font(bold=True)
        cell.alignment = _center()

    rows = [
        "EPS (Next Quarter)", "EPS (FY)", "Revenue (Next Quarter)",
        "Revenue (FY)", "EPS Growth (FY+1)", "Revenue Growth (FY+1)",
    ]
    for i, label in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).alignment = _left()
        for col in range(2, 7):
            ws.cell(row=i, column=col, value="—").alignment = _center()

    ws["A10"] = "⚠ Data populated by Yahoo Finance integration (Prompt 9)"
    ws["A10"].font = _font(color="808080")

    _freeze(ws)
    _autofit(ws)


# ── Sheet 5: Signals ───────────────────────────────────────────────────────

_LEVEL_CONFIG = {
    "critical": (_fill("FFC7CE"), _font(bold=True, color="9C0006"), "⚠ Critical"),
    "monitor":  (_fill("FFEB9C"), _font(bold=True, color="9C5700"), "△ Monitor"),
    "normal":   (_fill("C6EFCE"), _font(bold=True, color="276221"), "✓ Normal"),
}


def _build_signals(ws, signals: list[dict]) -> None:
    ws.title = "Signals"
    ws.sheet_properties.tabColor = _RED

    headers = ["Category", "Level", "Signal", "Explanation", "Note Reference", "Excerpt"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _fill(_RED)
        cell.font = _font(bold=True, color="FFFFFF")
        cell.alignment = _center()

    if not signals:
        ws.cell(row=2, column=1, value="No signals found.").font = _font(color="808080")
        _freeze(ws)
        _autofit(ws)
        return

    for i, sig in enumerate(signals, start=2):
        level  = sig.get("level", "normal").lower()
        bg, ft, label = _LEVEL_CONFIG.get(level, _LEVEL_CONFIG["normal"])

        ws.cell(row=i, column=1, value=sig.get("category_label", "")).alignment = _left()

        lc = ws.cell(row=i, column=2, value=label)
        lc.fill = bg; lc.font = ft; lc.alignment = _center()

        ws.cell(row=i, column=3, value=sig.get("title", "")).font = _font(bold=True)
        ws.cell(row=i, column=3).alignment = _left()

        ec = ws.cell(row=i, column=4, value=sig.get("explanation", ""))
        ec.alignment = Alignment(wrap_text=True, vertical="top")

        ws.cell(row=i, column=5, value=sig.get("note_reference", "")).alignment = _center()
        ws.cell(row=i, column=6, value=sig.get("excerpt", "")).alignment = \
            Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[i].height = 54

    _freeze(ws)
    _autofit(ws)
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["F"].width = 45
    ws.column_dimensions["C"].width = 35


# ── Public entry point ─────────────────────────────────────────────────────

def generate_workbook(
    ticker: str,
    year: str,
    income_df: pd.DataFrame,
    balance_df: pd.DataFrame,
    cashflow_df: pd.DataFrame,
    ratios: dict,
    signals: list[dict],
    filing_date: str = "",
) -> bytes:
    """
    Build the full 5-sheet Nuvel workbook and return raw bytes.
    Safe to call from a Celery task; no disk I/O inside this function.
    """
    wb = Workbook()

    # Remove the default empty sheet
    del wb[wb.sheetnames[0]]

    ws1 = wb.create_sheet("Raw Data")
    ws2 = wb.create_sheet("Ratios")
    ws3 = wb.create_sheet("DCF Model")
    ws4 = wb.create_sheet("Expectations")
    ws5 = wb.create_sheet("Signals")

    _build_raw_data(ws1, income_df, balance_df, cashflow_df)
    _build_ratios(ws2, ratios)
    _build_dcf(ws3, income_df, balance_df)
    _build_expectations(ws4, ticker, year)
    _build_signals(ws5, signals)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
